import pandas as pd
import numpy as np
from rapidfuzz import fuzz, process
import re
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class CrosswalkAnalyzer:
    """
    Core class for performing EnableNSW to NDIS crosswalk analysis
    """
    
    def __init__(self):
        self.confidence_threshold = 80
        self.include_repair_codes = True
        
        # Define mapping rules for known direct mappings
        self.mapping_rules = {
            # Personal Mobility
            'manual wheelchair': {
                'keywords': ['manual', 'wheelchair', 'push'],
                'ndis_category': 'Personal Mobility',
                'confidence': 'Direct line item (High confidence)'
            },
            'power wheelchair': {
                'keywords': ['power', 'electric', 'motorised', 'wheelchair'],
                'ndis_category': 'Personal Mobility', 
                'confidence': 'Direct line item (High confidence)'
            },
            'mobility scooter': {
                'keywords': ['scooter', 'mobility'],
                'ndis_category': 'Personal Mobility',
                'confidence': 'Direct line item (High confidence)'
            },
            'walking frame': {
                'keywords': ['walking frame', 'walker', 'rollator'],
                'ndis_category': 'Personal Mobility',
                'confidence': 'Direct line item (High confidence)'
            },
            
            # Communication
            'speech device': {
                'keywords': ['speech', 'communication', 'voice', 'aac'],
                'ndis_category': 'Communication',
                'confidence': 'Direct line item (High confidence)'
            },
            'hearing aid': {
                'keywords': ['hearing', 'audio', 'amplification'],
                'ndis_category': 'Hearing',
                'confidence': 'Direct line item (High confidence)'
            },
            
            # Vision
            'magnifier': {
                'keywords': ['magnify', 'vision', 'low vision', 'sight'],
                'ndis_category': 'Vision',
                'confidence': 'Direct line item (High confidence)'
            },
            'braille': {
                'keywords': ['braille', 'tactile'],
                'ndis_category': 'Vision',
                'confidence': 'Direct line item (High confidence)'
            },
            
            # Daily Living
            'bathroom aid': {
                'keywords': ['bathroom', 'toilet', 'shower', 'bath'],
                'ndis_category': 'Daily Living',
                'confidence': 'Best-fit (Functional equivalent)'
            },
            'kitchen aid': {
                'keywords': ['kitchen', 'cooking', 'dining'],
                'ndis_category': 'Daily Living',
                'confidence': 'Best-fit (Functional equivalent)'
            },
            
            # Seating and Positioning
            'cushion': {
                'keywords': ['cushion', 'seating', 'positioning'],
                'ndis_category': 'Seating and Positioning',
                'confidence': 'Direct line item (High confidence)'
            }
        }
    
    def clean_text(self, text):
        """Clean and normalize text for matching"""
        if pd.isna(text):
            return ""
        
        # Convert to string and lowercase
        text = str(text).lower()
        
        # Remove special characters but keep spaces and hyphens
        text = re.sub(r'[^\w\s\-]', ' ', text)
        
        # Replace multiple spaces with single space
        text = re.sub(r'\s+', ' ', text)
        
        return text.strip()
    
    def extract_keywords(self, text):
        """Extract meaningful keywords from text"""
        text = self.clean_text(text)
        
        # Common stop words to remove
        stop_words = {'the', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 'of', 'with', 'by', 'a', 'an'}
        
        words = text.split()
        keywords = [word for word in words if word not in stop_words and len(word) > 2]
        
        return keywords
    
    def validate_ndis_data(self, ndis_df):
        """Validate and clean NDIS data, handling various column name formats"""
        
        # Common column name variations for NDIS data
        column_mappings = {
            'Support_Item_Number': [
                'Support_Item_Number', 'Support Item Number', 'Item Number', 
                'Code', 'Support Code', 'NDIS Code', 'Item Code', 'Number'
            ],
            'Support_Item_Name': [
                'Support_Item_Name', 'Support Item Name', 'Item Name', 
                'Description', 'Support Item', 'Item Description', 'Name'
            ],
            'Category': [
                'Category', 'AT Category', 'Support Category', 'Type', 'Group'
            ],
            'Description': [
                'Description', 'Details', 'Item Description', 'Full Description'
            ],
            'Unit_Price': [
                'Unit_Price', 'Unit Price', 'Price', 'Cost', 'Amount'
            ]
        }
        
        # Create a copy to avoid modifying original
        df = ndis_df.copy()
        
        # Preserve Source_Table column if it exists
        source_table_col = None
        if 'Source_Table' in df.columns:
            source_table_col = df['Source_Table'].copy()
        
        # Standardize column names
        for standard_name, variations in column_mappings.items():
            for col in df.columns:
                if col.strip() in variations:
                    df = df.rename(columns={col: standard_name})
                    break
        
        # Check for required columns
        required_cols = ['Support_Item_Number', 'Support_Item_Name']
        missing_cols = [col for col in required_cols if col not in df.columns]
        
        if missing_cols:
            # Try to identify columns by content analysis
            df = self.identify_columns_by_content(df, missing_cols)
        
        # Ensure required columns exist
        for col in required_cols:
            if col not in df.columns:
                raise ValueError(f"Could not identify required column: {col}")
        
        # Clean data - remove rows where both required columns are empty
        df = df.dropna(subset=['Support_Item_Number', 'Support_Item_Name'], how='all')
        
        # Clean and standardize the data
        df['Support_Item_Number'] = df['Support_Item_Number'].astype(str).str.strip()
        df['Support_Item_Name'] = df['Support_Item_Name'].astype(str).str.strip()
        
        # Remove rows with empty or invalid codes/names
        df = df[df['Support_Item_Number'] != '']
        df = df[df['Support_Item_Number'] != 'nan']
        df = df[df['Support_Item_Name'] != '']
        df = df[df['Support_Item_Name'] != 'nan']
        
        # Add missing optional columns
        if 'Category' not in df.columns:
            df['Category'] = 'Unknown'
        if 'Description' not in df.columns:
            df['Description'] = df['Support_Item_Name']
        if 'Unit_Price' not in df.columns:
            df['Unit_Price'] = 0
        
        # Restore Source_Table column if it existed
        if source_table_col is not None:
            df['Source_Table'] = source_table_col
        
        return df
    
    def identify_columns_by_content(self, df, missing_cols):
        """Try to identify missing columns by analyzing content"""
        
        for col_name in missing_cols:
            best_match_col = None
            
            if col_name == 'Support_Item_Number':
                # Look for columns with code-like patterns
                for col in df.columns:
                    sample_values = df[col].dropna().astype(str).head(10)
                    code_pattern_count = sum(1 for val in sample_values 
                                           if re.match(r'^\d+_\d+.*', val) or 
                                              re.match(r'^\d{2,}', val))
                    
                    if code_pattern_count > len(sample_values) * 0.5:  # More than 50% match
                        best_match_col = col
                        break
            
            elif col_name == 'Support_Item_Name':
                # Look for columns with descriptive text
                for col in df.columns:
                    if col not in [best_match_col]:  # Don't reuse already identified columns
                        sample_values = df[col].dropna().astype(str).head(10)
                        avg_length = sum(len(val) for val in sample_values) / len(sample_values)
                        
                        if avg_length > 10:  # Descriptive text usually longer
                            best_match_col = col
                            break
            
            if best_match_col:
                df = df.rename(columns={best_match_col: col_name})
        
        return df
    
    def rule_based_matching(self, subcategory, description=""):
        """Apply rule-based matching for known patterns"""
        
        text_to_match = f"{subcategory} {description}".lower()
        
        for rule_name, rule_data in self.mapping_rules.items():
            # Check if any rule keywords are present
            keywords_found = []
            for keyword in rule_data['keywords']:
                if keyword in text_to_match:
                    keywords_found.append(keyword)
            
            # If we found keywords, this is likely a match
            if keywords_found:
                return {
                    'matched_rule': rule_name,
                    'ndis_category': rule_data['ndis_category'],
                    'confidence': rule_data['confidence'],
                    'keywords_matched': keywords_found,
                    'match_score': 95  # High score for rule-based matches
                }
        
        return None
    
    def fuzzy_matching(self, subcategory, ndis_items, description=""):
        """Perform fuzzy string matching against NDIS items"""
        
        text_to_match = f"{subcategory} {description}"
        
        # Create list of NDIS item descriptions for matching
        ndis_descriptions = []
        for idx, row in ndis_items.iterrows():
            item_text = f"{row.get('Support_Item_Name', '')} {row.get('Description', '')}"
            ndis_descriptions.append((self.clean_text(item_text), idx))
        
        if not ndis_descriptions:
            return None
        
        # Find best matches using rapidfuzz
        matches = process.extract(
            self.clean_text(text_to_match),
            [desc[0] for desc in ndis_descriptions],
            scorer=fuzz.WRatio,
            limit=3
        )
        
        best_matches = []
        for match_text, score, _ in matches:
            if score >= self.confidence_threshold:
                # Find the original index
                original_idx = next(idx for desc, idx in ndis_descriptions if desc == match_text)
                best_matches.append({
                    'ndis_index': original_idx,
                    'score': score,
                    'matched_text': match_text
                })
        
        return best_matches if best_matches else None
    
    def determine_confidence_level(self, match_score, rule_based=False):
        """Determine confidence level based on match score and method"""
        
        if rule_based:
            return "Direct line item (High confidence)"
        elif match_score >= 90:
            return "Direct line item (High confidence)"
        elif match_score >= 75:
            return "Best-fit (Functional equivalent)"
        else:
            return "No clear equivalent (Review required)"
    
    def run_crosswalk(self, enable_nsw_df, ndis_df):
        """Main function to run the crosswalk analysis"""
        
        # Clean and prepare data
        enable_nsw_df = enable_nsw_df.copy()
        
        # Validate and clean NDIS data (handles DOCX column variations)
        try:
            ndis_df = self.validate_ndis_data(ndis_df)
        except ValueError as e:
            raise ValueError(f"NDIS data validation failed: {str(e)}")
        
        # Ensure required columns exist in EnableNSW data
        if 'Description' not in enable_nsw_df.columns:
            enable_nsw_df['Description'] = ""
        
        # Initialize results list
        crosswalk_results = []
        
        # Process each EnableNSW subcategory
        for idx, row in enable_nsw_df.iterrows():
            category = row['Category']
            subcategory = row['Subcategory'] 
            description = row.get('Description', '')
            
            # Initialize result record
            result = {
                'EnableNSW_Category': category,
                'EnableNSW_Subcategory': subcategory,
                'EnableNSW_Description': description,
                'NDIS_Support_Item_Number': None,
                'NDIS_Support_Item_Name': None,
                'NDIS_Category': None,
                'NDIS_Description': None,
                'NDIS_Unit_Price': None,
                'NDIS_Source_Table': None,  # Track which table the match came from
                'Mapping_Confidence': 'No clear equivalent (Review required)',
                'Match_Score': 0,
                'Matching_Method': 'None',
                'Keywords_Matched': '',
                'Repair_Maintenance_Code': None
            }
            
            # Try rule-based matching first
            rule_match = self.rule_based_matching(subcategory, description)
            
            if rule_match:
                # Find NDIS items matching the rule category
                matching_ndis = ndis_df[
                    ndis_df['Category'].str.contains(rule_match['ndis_category'], case=False, na=False) |
                    ndis_df['Support_Item_Name'].str.contains('|'.join(rule_match['keywords_matched']), case=False, na=False)
                ]
                
                if not matching_ndis.empty:
                    # Take the first matching item
                    best_match = matching_ndis.iloc[0]
                    
                    result.update({
                        'NDIS_Support_Item_Number': best_match['Support_Item_Number'],
                        'NDIS_Support_Item_Name': best_match['Support_Item_Name'],
                        'NDIS_Category': best_match.get('Category', rule_match['ndis_category']),
                        'NDIS_Description': best_match.get('Description', ''),
                        'NDIS_Unit_Price': best_match.get('Unit_Price', 0),
                        'NDIS_Source_Table': best_match.get('Source_Table', 'Unknown'),
                        'Mapping_Confidence': rule_match['confidence'],
                        'Match_Score': rule_match['match_score'],
                        'Matching_Method': 'Rule-based',
                        'Keywords_Matched': ', '.join(rule_match['keywords_matched'])
                    })
                    
                    # Add repair/maintenance code if requested
                    if self.include_repair_codes:
                        repair_code = self.find_repair_code(best_match['Support_Item_Number'], ndis_df)
                        result['Repair_Maintenance_Code'] = repair_code
            
            # If no rule match, try fuzzy matching
            if result['NDIS_Support_Item_Number'] is None:
                fuzzy_matches = self.fuzzy_matching(subcategory, ndis_df, description)
                
                if fuzzy_matches:
                    # Take the best fuzzy match
                    best_fuzzy = fuzzy_matches[0]
                    best_match = ndis_df.iloc[best_fuzzy['ndis_index']]
                    
                    result.update({
                        'NDIS_Support_Item_Number': best_match['Support_Item_Number'],
                        'NDIS_Support_Item_Name': best_match['Support_Item_Name'], 
                        'NDIS_Category': best_match.get('Category', ''),
                        'NDIS_Description': best_match.get('Description', ''),
                        'NDIS_Unit_Price': best_match.get('Unit_Price', 0),
                        'NDIS_Source_Table': best_match.get('Source_Table', 'Unknown'),
                        'Mapping_Confidence': self.determine_confidence_level(best_fuzzy['score']),
                        'Match_Score': best_fuzzy['score'],
                        'Matching_Method': 'Fuzzy matching',
                        'Keywords_Matched': ', '.join(self.extract_keywords(subcategory)[:3])  # Top 3 keywords
                    })
                    
                    # Add repair/maintenance code if requested
                    if self.include_repair_codes:
                        repair_code = self.find_repair_code(best_match['Support_Item_Number'], ndis_df)
                        result['Repair_Maintenance_Code'] = repair_code
            
            crosswalk_results.append(result)
        
        # Convert to DataFrame
        crosswalk_df = pd.DataFrame(crosswalk_results)
        
        # Generate pivot summary
        pivot_summary = self.generate_pivot_summary(crosswalk_df)
        
        return {
            'crosswalk': crosswalk_df,
            'pivot_summary': pivot_summary,
            'metadata': {
                'total_items': len(crosswalk_df),
                'mapped_items': len(crosswalk_df[crosswalk_df['NDIS_Support_Item_Number'].notna()]),
                'analysis_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'confidence_threshold': self.confidence_threshold
            }
        }
    
    def find_repair_code(self, support_item_number, ndis_df):
        """Find associated repair/maintenance code for a support item"""
        
        if pd.isna(support_item_number):
            return None
        
        # Look for repair codes with similar patterns
        repair_patterns = ['repair', 'maintenance', 'service']
        
        for pattern in repair_patterns:
            repair_items = ndis_df[
                ndis_df['Support_Item_Name'].str.contains(pattern, case=False, na=False)
            ]
            
            if not repair_items.empty:
                # Return the first repair code found
                return repair_items.iloc[0]['Support_Item_Number']
        
        return None
    
    def generate_pivot_summary(self, crosswalk_df):
        """Generate pivot table summary of mappings"""
        
        # Count mappings by EnableNSW Category and NDIS Category
        pivot_data = []
        
        for enable_category in crosswalk_df['EnableNSW_Category'].unique():
            category_data = crosswalk_df[crosswalk_df['EnableNSW_Category'] == enable_category]
            
            # Count total subcategories
            total_subcategories = len(category_data)
            
            # Count mapped items
            mapped_items = len(category_data[category_data['NDIS_Support_Item_Number'].notna()])
            
            # Count by confidence level
            confidence_counts = category_data['Mapping_Confidence'].value_counts().to_dict()
            
            # Get NDIS categories mapped to
            ndis_categories = category_data[category_data['NDIS_Category'].notna()]['NDIS_Category'].unique()
            
            pivot_data.append({
                'EnableNSW_Category': enable_category,
                'Total_Subcategories': total_subcategories,
                'Mapped_Items': mapped_items,
                'Mapping_Rate': f"{mapped_items/total_subcategories*100:.1f}%",
                'NDIS_Categories': ', '.join(ndis_categories) if len(ndis_categories) > 0 else 'None',
                'High_Confidence': confidence_counts.get('Direct line item (High confidence)', 0),
                'Best_Fit': confidence_counts.get('Best-fit (Functional equivalent)', 0),
                'Review_Required': confidence_counts.get('No clear equivalent (Review required)', 0)
            })
        
        return pd.DataFrame(pivot_data)
    
    def generate_excel_report(self, results):
        """Generate comprehensive Excel report with multiple sheets"""
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Introduction sheet
        intro_ws = wb.create_sheet("Introduction")
        self.create_introduction_sheet(intro_ws, results['metadata'])
        
        # Create Crosswalk Table sheet
        crosswalk_ws = wb.create_sheet("Crosswalk Table")
        self.create_crosswalk_sheet(crosswalk_ws, results['crosswalk'])
        
        # Create Pivot Summary sheet
        pivot_ws = wb.create_sheet("Pivot Summary")
        self.create_pivot_sheet(pivot_ws, results['pivot_summary'])
        
        # Save to buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
    
    def create_introduction_sheet(self, worksheet, metadata):
        """Create introduction and methodology sheet"""
        
        # Title
        worksheet['A1'] = "EnableNSW to NDIS Crosswalk Analysis Report"
        worksheet['A1'].font = Font(size=16, bold=True)
        
        # Metadata
        row = 3
        worksheet[f'A{row}'] = "Analysis Date:"
        worksheet[f'B{row}'] = metadata['analysis_date']
        
        row += 1
        worksheet[f'A{row}'] = "Total Items Analyzed:"
        worksheet[f'B{row}'] = metadata['total_items']
        
        row += 1  
        worksheet[f'A{row}'] = "Successfully Mapped:"
        worksheet[f'B{row}'] = metadata['mapped_items']
        
        row += 1
        worksheet[f'A{row}'] = "Mapping Success Rate:"
        worksheet[f'B{row}'] = f"{metadata['mapped_items']/metadata['total_items']*100:.1f}%"
        
        # Methodology
        row += 3
        worksheet[f'A{row}'] = "METHODOLOGY"
        worksheet[f'A{row}'].font = Font(size=14, bold=True)
        
        methodology_text = [
            "",
            "This crosswalk analysis uses a two-stage matching approach:",
            "",
            "1. RULE-BASED MATCHING",
            "   - Direct mappings for known equipment categories",
            "   - High confidence matches based on predefined rules",
            "   - Examples: manual wheelchair → NDIS personal mobility codes",
            "",
            "2. FUZZY STRING MATCHING", 
            "   - Approximate matching using RapidFuzz library",
            "   - Confidence threshold: " + str(metadata['confidence_threshold']) + "%",
            "   - Best-fit matches for similar descriptions",
            "",
            "CONFIDENCE LEVELS:",
            "",
            "• Direct line item (High confidence)",
            "  - Exact or rule-based matches",
            "  - Match score ≥ 90%",
            "",
            "• Best-fit (Functional equivalent)",
            "  - Good fuzzy matches", 
            "  - Match score 75-89%",
            "",
            "• No clear equivalent (Review required)",
            "  - Low confidence or no matches found",
            "  - Match score < 75% or no match",
            "",
            "LIMITATIONS:",
            "",
            "• Automated matching may miss contextual nuances",
            "• Manual review recommended for 'Review required' items",
            "• NDIS pricing and availability subject to change"
        ]
        
        for text in methodology_text:
            row += 1
            worksheet[f'A{row}'] = text
    
    def create_crosswalk_sheet(self, worksheet, crosswalk_df):
        """Create crosswalk table sheet"""
        
        # Add title
        worksheet['A1'] = "EnableNSW to NDIS Crosswalk Mapping"
        worksheet['A1'].font = Font(size=14, bold=True)
        
        # Add data starting from row 3
        for r_idx, row in enumerate(dataframe_to_rows(crosswalk_df, index=False, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                
                # Style header row
                if r_idx == 3:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                # Color code confidence levels
                elif r_idx > 3 and c_idx == crosswalk_df.columns.get_loc('Mapping_Confidence') + 1:
                    if value == "Direct line item (High confidence)":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value == "Best-fit (Functional equivalent)":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif value == "No clear equivalent (Review required)":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_pivot_sheet(self, worksheet, pivot_df):
        """Create pivot summary sheet"""
        
        # Add title
        worksheet['A1'] = "EnableNSW to NDIS Mapping Summary"
        worksheet['A1'].font = Font(size=14, bold=True)
        
        # Add data starting from row 3
        for r_idx, row in enumerate(dataframe_to_rows(pivot_df, index=False, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                
                # Style header row
                if r_idx == 3:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width